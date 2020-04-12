import openpyxl
import os

from configuration.config import configuration
from configuration.path import EXCEL_FILE_PATH


class CaseData(object):
    pass


class HandleExcel(object):
    def __init__(self, sheetname=None, filename=None):
        if filename is None:
            self.filename = os.path.join(EXCEL_FILE_PATH, configuration.getConfig('excel', 'excelname'))
        else:
            self.filename = filename

        self.sheetname = sheetname

    def load(self):
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
            if self.sheetname is None:
                self.sheet = self.workbook[self.workbook.sheetnames[0]]
            else:
                self.sheet = self.workbook[self.sheetname]
        except Exception:
            print("open excel failed")

    def read_data_to_obj(self):
        """
        加载excel用例到CaseData类
        :return: CaseData类列表
        """
        self.load()
        # 按行获取表格中的数据，将获取出来的数据放在一个列表里面
        try:
            rows = list(self.sheet.rows)
            cases = []

            head_list = [row.value for row in rows[0]]

            # 获取表头以外的数据
            for row_item in rows[1:]:
                content_list = [i.value for i in row_item]

                case_data = CaseData()
                for i in zip(head_list, content_list):
                    setattr(case_data, i[0], i[1])  # 将每行数据转化为一个实例对象
                cases.append(case_data)
            return cases
        except Exception:
            print("read file to obj failed")
        finally:
            self.workbook.close()

    def write_to_file(self, row, column, value):
        """
        将数据写入excel
        :param row: 行号
        :param column: 列号
        :param value: 写入的值
        :return: None
        """
        self.load()
        try:
            # 往文件中写入内容
            self.sheet.cell(row=row, column=column, value=value)
            # 保存文件内容
            self.workbook.save(self.filename)
            # 关闭文件
        except Exception:
            print("fail to write to excel")
        finally:
            self.workbook.close()


if __name__ == '__main__':
    # data = HandleExcel('test_set.xlsx', 'register')
    # data.write_to_file(1, 4, 'result')
    data = HandleExcel()
    print(data.read_data_to_obj()[0].__getattribute__("url"))
    pass

